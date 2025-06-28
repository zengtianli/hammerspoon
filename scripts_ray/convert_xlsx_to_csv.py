#!/usr/bin/env python3
"""
XLSX转CSV转换工具 - 将Excel XLSX文件转换为CSV格式
版本: 2.0.0
作者: tianli
更新: 2024-01-01
"""

import sys
import csv
import argparse
from pathlib import Path
from typing import Optional, List

# 引入通用工具模块
from common_utils import (
    show_success, show_error, show_warning, show_info, show_processing,
    validate_input_file, check_file_extension, get_file_basename,
    ProgressTracker, fatal_error, check_python_packages,
    show_version_info, find_files_by_extension
)

# 脚本版本信息
SCRIPT_VERSION = "2.0.0"
SCRIPT_AUTHOR = "tianli"
SCRIPT_UPDATED = "2024-01-01"

def check_dependencies() -> bool:
    """检查依赖"""
    show_info("检查依赖项...")
    
    # 检查必要的Python包
    if not check_python_packages(['openpyxl']):
        return False
    
    show_success("依赖检查完成")
    return True

def convert_xlsx_to_csv_single(input_file: Path, output_file: Optional[Path] = None, 
                              sheet_name: Optional[str] = None, 
                              all_sheets: bool = True) -> bool:
    """将单个XLSX文件转换为CSV格式"""
    try:
        # 验证输入文件
        if not validate_input_file(input_file):
            return False
        
        # 检查文件扩展名
        if not check_file_extension(input_file, 'xlsx'):
            show_warning(f"跳过不支持的文件: {input_file.name}")
            return False
        
        show_processing(f"转换: {input_file.name}")
        
        # 导入openpyxl
        try:
            from openpyxl import load_workbook
        except ImportError:
            show_error("缺少依赖包: openpyxl")
            return False
        
        wb = load_workbook(input_file, read_only=True, data_only=True)
        
        # 处理工作表
        if sheet_name and sheet_name in wb.sheetnames:
            # 处理指定工作表
            sheets_to_process = [(sheet_name, wb[sheet_name])]
        elif not all_sheets:
            # 只处理活动工作表
            sheets_to_process = [('active', wb.active)]
        else:
            # 处理所有工作表
            sheets_to_process = [(name, wb[name]) for name in wb.sheetnames]
        
        success_count = 0
        for sheet_name, ws in sheets_to_process:
            # 生成输出文件名
            if output_file and len(sheets_to_process) == 1:
                current_output = output_file
            else:
                if len(sheets_to_process) > 1:
                    current_output = input_file.parent / f"{input_file.stem}_{sheet_name}.csv"
                else:
                    current_output = input_file.with_suffix('.csv')
            
            # 写入CSV文件
            with open(current_output, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    # 将None转换为空字符串
                    writer.writerow(['' if cell is None else str(cell) for cell in row])
            
            show_success(f"已转换工作表 '{sheet_name}' -> {current_output.name}")
            success_count += 1
        
        return success_count > 0
        
    except Exception as e:
        show_error(f"转换失败: {input_file.name} - {e}")
        return False

def batch_process(directory: Path, recursive: bool = False, all_sheets: bool = True) -> None:
    """批量处理XLSX文件"""
    show_info(f"处理目录: {directory}")
    
    # 查找XLSX文件
    files = find_files_by_extension(directory, 'xlsx', recursive)
    
    if not files:
        show_warning("未找到XLSX文件")
        return
    
    show_info(f"找到 {len(files)} 个XLSX文件")
    
    # 初始化进度跟踪器
    tracker = ProgressTracker()
    
    # 处理每个文件
    for i, file in enumerate(files, 1):
        show_progress(i, len(files), file.name)
        
        if convert_xlsx_to_csv_single(file, all_sheets=all_sheets):
            tracker.add_success()
        else:
            tracker.add_failure()
    
    # 显示统计
    tracker.show_summary("文件转换")

def show_version() -> None:
    """显示版本信息"""
    show_version_info(SCRIPT_VERSION, SCRIPT_AUTHOR, SCRIPT_UPDATED)

def show_help() -> None:
    """显示帮助信息"""
    print(f"""
XLSX转CSV转换工具 - 将Excel XLSX文件转换为CSV格式

用法:
    python3 {sys.argv[0]} [选项] [输入] [输出]

参数:
    输入            输入XLSX文件或目录
    输出            输出CSV文件（可选，仅对单文件有效）

选项:
    -r, --recursive   递归处理子目录
    -s, --sheet NAME  指定要转换的工作表名称
    -d, --default     仅转换默认工作表（不处理所有工作表）
    -h, --help        显示此帮助信息
    --version         显示版本信息

示例:
    python3 {sys.argv[0]} data.xlsx                    # 转换所有工作表
    python3 {sys.argv[0]} data.xlsx output.csv        # 指定输出文件
    python3 {sys.argv[0]} -s Sheet1 data.xlsx         # 仅转换指定工作表
    python3 {sys.argv[0]} -d data.xlsx                # 仅转换默认工作表
    python3 {sys.argv[0]} ./data_dir                   # 批量转换目录
    python3 {sys.argv[0]} -r ./data_dir                # 递归转换目录

功能:
    - 将Excel XLSX文件转换为CSV格式
    - 支持转换所有工作表或指定工作表
    - 支持单文件和批量处理
    - 自动处理编码问题

依赖:
    - openpyxl
    """)

def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='XLSX转CSV转换工具',
        add_help=False
    )
    
    parser.add_argument('input', nargs='?', help='输入XLSX文件或目录')
    parser.add_argument('output', nargs='?', help='输出CSV文件')
    parser.add_argument('-r', '--recursive', action='store_true', help='递归处理子目录')
    parser.add_argument('-s', '--sheet', help='指定要转换的工作表名称')
    parser.add_argument('-d', '--default', action='store_true', help='仅转换默认工作表')
    parser.add_argument('-h', '--help', action='store_true', help='显示帮助信息')
    parser.add_argument('--version', action='store_true', help='显示版本信息')
    
    args = parser.parse_args()
    
    if args.help:
        show_help()
        return
    
    if args.version:
        show_version()
        return
    
    # 检查依赖
    if not check_dependencies():
        sys.exit(1)
    
    # 处理输入
    if not args.input:
        # 默认处理当前目录
        batch_process(Path.cwd(), all_sheets=not args.default)
    else:
        input_path = Path(args.input)
        
        if input_path.is_file():
            # 单文件处理
            output_path = None
            if args.output:
                output_path = Path(args.output)
            
            # 确定是否处理所有工作表
            all_sheets = not (args.sheet or args.default)
            
            if not convert_xlsx_to_csv_single(input_path, output_path, 
                                            args.sheet, all_sheets):
                sys.exit(1)
        elif input_path.is_dir():
            # 目录处理
            batch_process(input_path, args.recursive, all_sheets=not args.default)
        else:
            fatal_error(f"输入路径不存在: {input_path}")

if __name__ == "__main__":
    main() 